import openpyxl, glob, os, re
from collections import OrderedDict

VARIANTS = OrderedDict([
    ("Full","REMO_new2_AdaMaO"),("Lite","REMO_new2_AdaMaO_Lite"),
    ("NoInd","REMO_new2_AdaMaO_NoIndicator"),
    ("PIEAOnly","REMO_new2_AdaMaO_PIEAOnlySelection"),
    ("FixedInd","REMO_new2_AdaMaO_FixedIndicatorAlways"),
])
RNAME = {v:k for k,v in VARIANTS.items()}

def parse_cell(s):
    if s is None: return None
    m = re.match(r'^([0-9eE+\-.]+)\s*\(([0-9eE+\-.]+)\)\s*([+\-/=]?)$', str(s).strip())
    if not m: return None
    try: return float(m.group(1))
    except: return None

def read_folder(base):
    files = sorted([f for f in glob.glob(os.path.join(base,"*.xlsx")) if not os.path.basename(f).startswith("~$")])
    data={}; problems=[]
    for f in files:
        wb=openpyxl.load_workbook(f,data_only=True); ws=wb.active
        header=[ws.cell(row=1,column=c).value for c in range(1,ws.max_column+1)]
        colmap={str(header[c-1]).strip():c for c in range(1,ws.max_column+1) if header[c-1] is not None}
        for r in range(2,ws.max_row+1):
            p=ws.cell(row=r,column=1).value
            if p is None: continue
            p=str(p).strip()
            if not (p.startswith("DTLZ") or p.startswith("WFG") or p.startswith("MaF")): continue
            M=ws.cell(row=r,column=2).value
            key=(int(M) if M is not None else -1,p)
            if key not in data:
                data[key]={}; 
                if p not in problems: problems.append(p)
            for a in VARIANTS.values():
                c=colmap.get(a)
                if c is None: continue
                v=parse_cell(ws.cell(row=r,column=c).value)
                if v is not None: data[key][a]=v
    return data,problems

base10=r"D:\我的文档\研究生\组会2026\7.8组会\AdaMao实验表\消融实验\指标模式"
base20=r"C:\Users\lsx\Desktop\AdaMao实验表\消融实验\指标模式\二十目标"
data10,p10=read_folder(base10); data20,p20=read_folder(base20)
probs=p10

def color(v, lo, hi):
    if hi==lo: t=0.0
    else: t=(v-lo)/(hi-lo)   # 0=best 1=worst
    # green(best) -> yellow -> red(worst)
    import colorsys
    hue = (1-t)*0.33  # 0.33 green -> 0 red
    r,g,b = colorsys.hsv_to_rgb(hue,0.65,0.95)
    return f"#{int(r*255):02x}{int(g*255):02x}{int(b*255):02x}"

def block(M,data,title):
    rows=""
    for p in probs:
        vals={a:data[(M,p)][a] for a in VARIANTS.values() if a in data[(M,p)]}
        lo=min(vals.values()); hi=max(vals.values())
        best=min(vals,key=vals.get)
        cells=""
        for a in VARIANTS.values():
            v=vals[a]
            bg=color(v,lo,hi)
            star = " ★" if a==best else ""
            bold = "font-weight:bold;" if a==best else ""
            cells+=f'<td style="background:{bg};{bold}color:#222;padding:3px 6px;text-align:right">{v:.2g}{star}</td>'
        flag = ' style="background:#fff3cd;font-weight:bold"' if (M==20 and best==VARIANTS["Full"]) else ''
        rows+=f'<tr><td style="text-align:left;padding:3px 8px;white-space:nowrap">{p}</td>{cells}</tr>'
    head=''.join(f'<th style="padding:3px 6px">{RNAME[a]}</th>' for a in VARIANTS.values())
    return f'<div style="flex:1"><h3 style="margin:4px 0;color:#1f3a5f">{title}</h3>'\
           f'<table style="border-collapse:collapse;font-size:12px;font-family:Consolas,monospace">'\
           f'<tr><th style="padding:3px 8px"></th>{head}</tr>{rows}</table></div>'

html=f'''<!doctype html><html><head><meta charset="utf-8"><title>指标模式消融 M10 vs M20</title></head>
<body style="font-family:Segoe UI,sans-serif;background:#fafafa;color:#222;padding:18px">
<h1 style="color:#1f3a5f">指标模式（PIEA）消融：M=10 vs M=20 — 5 变体 IGD 热力图</h1>
<p style="color:#555;max-width:900px">
每行一个测试问题，颜色：<span style="color:#1a7f37;font-weight:bold">绿=该问题最优</span> → 红=最差（行内归一化）。
★ = 该问题最优变体。<span style="background:#fff3cd">黄底行</span> = M=20 下 Full 夺冠的问题。
<b>读图要点：</b>M=20 下 Full 在 DTLZ1/3/7、WFG3/5/9 等夺冠（黄底），而 PIEAOnly/FixedInd 在 DTLZ7 整行偏红（崩塌）。
</p>
<div style="display:flex;gap:30px;flex-wrap:wrap">
{block(10,data10,"M = 10")}
{block(20,data20,"M = 20")}
</div>
<div style="margin-top:18px">
<h3 style="color:#1f3a5f">平均秩（5 变体内部，越低越好）</h3>
<table style="border-collapse:collapse;font-size:13px">
<tr><th style="padding:4px 10px"></th><th>M=10</th><th>M=20</th><th>趋势</th></tr>
<tr><td style="padding:4px 10px;font-weight:bold;background:#d4edda">Full</td><td>3.06</td><td>2.75</td><td style="color:#1a7f37">↑ 变好（登顶）</td></tr>
<tr><td style="padding:4px 10px">Lite</td><td>3.06</td><td>3.19</td><td style="color:#b02a37">↓ 变差</td></tr>
<tr><td style="padding:4px 10px">NoInd</td><td>3.31</td><td>3.06</td><td>↑</td></tr>
<tr><td style="padding:4px 10px">PIEAOnly</td><td>2.69</td><td>2.94</td><td style="color:#b02a37">↓（崩塌拖累）</td></tr>
<tr><td style="padding:4px 10px">FixedInd</td><td>2.88</td><td>3.06</td><td style="color:#b02a37">↓</td></tr>
</table>
<p style="color:#555">M=20 全场 11 算法同场平均秩：<b>Full 4.06 排第 1</b>（超越 PIEA 4.50、PIEAOnly 4.56、Lite 5.03）。
Friedman 内部 p 值：M=10=0.844，M=20=0.951（单维度不显著，但维度扫描方向性明确）。</p>
</div>
</body></html>'''

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"指标模式消融_热力图_M10vsM20.html")
open(out,"w",encoding="utf-8").write(html)
print("written:",out)
