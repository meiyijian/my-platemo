import openpyxl, glob, os, re
from collections import OrderedDict

VARIANTS = OrderedDict([
    ("Full",      "REMO_new2_AdaMaO"),
    ("Lite",      "REMO_new2_AdaMaO_Lite"),
    ("NoInd",     "REMO_new2_AdaMaO_NoIndicator"),
    ("PIEAOnly",  "REMO_new2_AdaMaO_PIEAOnlySelection"),
    ("FixedInd",  "REMO_new2_AdaMaO_FixedIndicatorAlways"),
])
RNAME = {v:k for k,v in VARIANTS.items()}
BASELINES = ["REMO_new2", "REMO", "PIEA", "MCEAD"]

def parse_cell(s):
    if s is None: return None
    s = str(s).strip()
    if s == '': return None
    m = re.match(r'^([0-9eE+\-.]+)\s*\(([0-9eE+\-.]+)\)\s*([+\-/=]?)$', s)
    if not m: return None
    try:
        return (float(m.group(1)), float(m.group(2)), m.group(3))
    except: return None

def read_folder(base):
    files = sorted([f for f in glob.glob(os.path.join(base,"*.xlsx")) if not os.path.basename(f).startswith("~$")])
    data = {}; problems = []
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True); ws = wb.active
        header = [ws.cell(row=1,column=c).value for c in range(1, ws.max_column+1)]
        colmap = {str(header[c-1]).strip(): c for c in range(1, ws.max_column+1) if header[c-1] is not None}
        for r in range(2, ws.max_row+1):
            pname = ws.cell(row=r, column=1).value
            if pname is None: continue
            pname = str(pname).strip()
            if not (pname.startswith("DTLZ") or pname.startswith("WFG") or pname.startswith("MaF")):
                continue
            M = ws.cell(row=r, column=2).value
            D = ws.cell(row=r, column=3).value
            key = (int(M) if M is not None else -1, pname)
            if key not in data:
                data[key] = {"_D": D}
                if pname not in problems: problems.append(pname)
            for algo in list(VARIANTS.values()) + BASELINES:
                c = colmap.get(algo)
                if c is None: continue
                pv = parse_cell(ws.cell(row=r, column=c).value)
                if pv is not None: data[key][algo] = pv
    return data, problems

def friedman(data, M, problems, algos):
    k = len(algos); N = len(problems)
    Rj = {a:0.0 for a in algos}
    for p in problems:
        vals = [(a, data[(M,p)][a][0]) for a in algos if a in data[(M,p)]]
        order = sorted(range(len(vals)), key=lambda i: vals[i][1])
        ranks = [0.0]*len(vals); i = 0
        while i < len(vals):
            j = i
            while j < len(vals)-1 and vals[order[j+1]][1] == vals[order[i]][1]: j += 1
            avg = (i+j)/2 + 1
            for t in range(i, j+1): ranks[order[t]] = avg
            i = j+1
        for idx,(a,val) in enumerate(vals): Rj[a] += ranks[idx]
    chi2 = 12.0/(N*k*(k+1)) * sum(Rj[a]**2 for a in algos) - 3*N*(k+1)
    df = k-1; pval = None
    try:
        from scipy import stats
        pval = float(stats.chi2.sf(chi2, df))
    except Exception: pass
    return chi2, df, pval, {a: Rj[a]/N for a in algos}

base10 = r"D:\我的文档\研究生\组会2026\7.8组会\AdaMao实验表\消融实验\指标模式"
base20 = r"C:\Users\lsx\Desktop\AdaMao实验表\消融实验\指标模式\二十目标"
data10, prob10 = read_folder(base10)
data20, prob20 = read_folder(base20)
print("M=10 problems:", len(prob10), "| M=20 problems:", len(prob20))

for M, data, problems, label in [(10,data10,prob10,"M=10"), (20,data20,prob20,"M=20")]:
    print("\n" + "#"*90)
    print(f"### {label} ({len(problems)} problems) — 5 AdaMaO variants ###")
    print("#"*90)
    chi2, df, pval, ranks = friedman(data, M, problems, list(VARIANTS.values()))
    print(f"Friedman: chi2={chi2:.4f} df={df} p={pval if pval is not None else 'n/a'}")
    print("Avg rank (lower=better):")
    for a in VARIANTS.values(): print(f"  {RNAME[a]:10} {ranks[a]:.3f}")
    best = {a:0 for a in VARIANTS.values()}; worst = {a:0 for a in VARIANTS.values()}
    for p in problems:
        vals = {a: data[(M,p)][a][0] for a in VARIANTS.values() if a in data[(M,p)]}
        if not vals: continue
        best[min(vals,key=vals.get)] += 1; worst[max(vals,key=vals.get)] += 1
    print("#best :", {RNAME[a]:best[a] for a in VARIANTS.values()})
    print("#worst:", {RNAME[a]:worst[a] for a in VARIANTS.values()})

print("\n" + "="*90)
print("CROSS-DIM avg rank per variant:")
for M, data, problems, label in [(10,data10,prob10,"M=10"), (20,data20,prob20,"M=20")]:
    _,_,_,ranks = friedman(data, M, problems, list(VARIANTS.values()))
    print(f"  {label}: " + "  ".join(f"{RNAME[a]}={ranks[a]:.2f}" for a in VARIANTS.values()))

print("\n" + "="*90)
print("FULL vs LITE vs PIEAOnly vs FixedInd vs NoInd — per problem (M10 / M20):")
print("="*90)
for p in prob10:
    r10 = {a: data10[(10,p)][a][0] for a in VARIANTS.values() if a in data10[(10,p)]}
    r20 = {a: data20[(20,p)][a][0] for a in VARIANTS.values() if a in data20[(20,p)]}
    f10 = "  ".join(f"{RNAME[a]}={r10[a]:.3g}" for a in VARIANTS.values())
    f20 = "  ".join(f"{RNAME[a]}={r20[a]:.3g}" for a in VARIANTS.values())
    # mark which variant is best at M20
    b20 = min(r20, key=r20.get)
    print(f"{p:8} M10: {f10}")
    print(f"{'':8} M20: {f20}   -> best@M20={RNAME[b20]}")

# ---- Full 11-algorithm rank (5 variants + 4 baselines) ----
print("\n" + "="*90)
print("FULL 11-algorithm avg rank (5 AdaMaO variants + 4 baselines):")
print("="*90)
ALL = list(VARIANTS.values()) + BASELINES
for M, data, problems, label in [(10,data10,prob10,"M=10"), (20,data20,prob20,"M=20")]:
    _,_,_,ranks = friedman(data, M, problems, ALL)
    order = sorted(ALL, key=lambda a: ranks[a])
    print(f"\n  {label} (lower=better):")
    for i,a in enumerate(order,1):
        tag = RNAME.get(a, a)
        print(f"    {i:2}. {tag:10} {ranks[a]:.2f}")

