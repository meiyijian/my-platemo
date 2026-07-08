import openpyxl, glob, os, re, math
from statistics import mean

base = r"D:\我的文档\研究生\组会2026\7.8组会\AdaMao实验表\消融实验\指标模式"
files = sorted([f for f in glob.glob(os.path.join(base, "*.xlsx")) if not os.path.basename(f).startswith("~$")])

# map variant column headers to short names
variant_map = {
    'REMO_new2_AdaMaO': 'Full',
    'REMO_new2_AdaMaO_NoIndicator': 'NoIndicator',
    'REMO_new2_AdaMaO_Lite': 'Lite',
    'REMO_new2_AdaMaO_PIEAOnlySelection': 'PIEAonly',
    'REMO_new2_AdaMaO_FixedIndicatorAlways': 'FixedIndicator',
}
baseline_map = {
    'REMO_new2': 'REMO_new2',
    'REMO': 'REMO',
    'PIEA': 'PIEA',
    'MCEAD': 'MCEAD',
    'REMO_new2_WFG10': 'WFG10',
    'R2AEA': 'R2AEA',
}

pat = re.compile(r'([-+]?\d+\.\d+e[+-]\d+)\s*\(([-+]?\d+\.\d+e[+-]\d+)\)\s*([+\-=])?')

# store: prob -> algo -> (mean, std, sign_in_file_with_algo_as_control)
data = {}   # data[prob][algo] = mean
data_std = {}
all_algos = set()
problems = []

for f in files:
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    for r in range(2, ws.max_row+1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
        prob = row[0]
        if not isinstance(prob, str) or not (prob.startswith('DTLZ') or prob.startswith('WFG')):
            continue
        problems.append(prob)
        for ci, name in enumerate(header):
            if name is None:
                continue
            val = row[ci]
            if not isinstance(val, str):
                continue
            m = pat.search(val)
            if not m:
                continue
            mean_v = float(m.group(1))
            std_v = float(m.group(2))
            data.setdefault(prob, {})[name] = mean_v
            data_std.setdefault(prob, {})[name] = std_v
            all_algos.add(name)
problems = list(dict.fromkeys(problems))

variants = ['Full','NoIndicator','Lite','PIEAonly','FixedIndicator']
# verify consistency across files for variant values (they should be identical)
print("Problems:", len(problems))

# 1) Average rank per problem across the 5 variants + #best
print("\n=== Per-problem IGD means (variants) ===")
print(f"{'Problem':8}" + "".join(f"{v:>13}" for v in variants) + "   best")
for p in problems:
    row = [data[p][vm] for vm in variant_map if vm in data[p]]
    # need actual keys
    vals = []
    for vkey in variant_map:
        vals.append(data[p].get(vkey))
    best = min(vals)
    best_v = [v for v, x in zip(variants, vals) if x==best]
    print(f"{p:8}" + "".join(f"{x:13.4g}" for x in vals) + "   " + "/".join(best_v))

# ranks
print("\n=== Average rank (lower better) & #best across 16 problems ===")
rank_sums = {v:0 for v in variants}
best_count = {v:0 for v in variants}
for p in problems:
    vals = [data[p][vk] for vk in variant_map]
    # rank: 1 = best (smallest); ties get average rank
    order = sorted(range(5), key=lambda i: vals[i])
    ranks = [0]*5
    i = 0
    while i < 5:
        j = i
        while j < 4 and vals[order[j+1]] == vals[order[i]]:
            j += 1
        avg_r = (i + j)/2 + 1
        for k in range(i, j+1):
            ranks[order[k]] = avg_r
        i = j+1
    for idx, v in enumerate(variants):
        rank_sums[v] += ranks[idx]
    mn = min(vals)
    for idx, x in enumerate(vals):
        if x == mn:
            best_count[variants[idx]] += 1
for v in variants:
    print(f"{v:13} avg_rank={rank_sums[v]/len(problems):.3f}  #best={best_count[v]}")

# 2) Friedman test across 5 variants, 16 problems (block=problem)
print("\n=== Friedman test (5 variants x 16 problems) ===")
N = len(problems)  # blocks
k = 5
Rj = {v: rank_sums[v] for v in variants}
chi2 = 12.0/(N*k*(k+1)) * sum(Rj[v]**2 for v in variants) - 3*N*(k+1)
df = k-1
# p-value for df=4: SF(x) = exp(-x/2)*(1 + x/2)   (exact closed form)
if df == 4:
    pval = math.exp(-chi2/2.0)*(1 + chi2/2.0)
else:
    pval = float('nan')
print(f"chi2={chi2:.4f}  df={df}  p={pval:.4f}")

# 3) Pairwise Full vs each variant: approximate z (n=30)
print("\n=== Full vs each: approx z-test (n=30), + = Full better (lower IGD) ===")
n = 30
for v in ['NoIndicator','Lite','PIEAonly','FixedIndicator']:
    full_key = 'REMO_new2_AdaMaO'; vkey = [k2 for k2,sv in variant_map.items() if sv==v][0]
    wins_full=wins_other=tie=0
    z_list=[]
    for p in problems:
        m1 = data[p][full_key]; s1 = data_std[p][full_key]
        m2 = data[p][vkey]; s2 = data_std[p][vkey]
        se = math.sqrt(s1**2/n + s2**2/n)
        if se == 0: 
            z=0
        else:
            z = (m1-m2)/se
        z_list.append(z)
        if abs(z) > 1.96:
            if m1 < m2: wins_full += 1
            else: wins_other += 1
        else:
            tie += 1
    print(f"Full vs {v:13}: Full-better={wins_full}  {v}-better={wins_other}  no-diff={tie}")

# 4) Baseline win counts
print("\n=== AdaMaO variants vs baselines: #problems with LOWER IGD (win) ===")
print(f"{'variant':13}" + "".join(f"{b:>10}" for b in baseline_map.values()))
for v in variants:
    vkey = [k2 for k2,sv in variant_map.items() if sv==v][0]
    row=[]
    for bkey in baseline_map:
        w=sum(1 for p in problems if data[p][vkey] < data[p][bkey])
        row.append(w)
    print(f"{v:13}" + "".join(f"{w:10}" for w in row))

# 5) Disconnected/degenerate PF spotlight (DTLZ7, WFG6) and connected regular (DTLZ2,3,5)
print("\n=== Spotlight: connected-regular vs disconnected/degenerate PFs ===")
connected = ['DTLZ1','DTLZ2','DTLZ3','DTLZ4','DTLZ5','WFG2','WFG9']
disconn = ['DTLZ7','WFG6']
for grp, lst in [('CONNECTED-regular', connected), ('DISCONNECTED/degenerate', disconn)]:
    print(f"\n-- {grp} --")
    print(f"{'Problem':8}" + "".join(f"{v:>13}" for v in variants))
    for p in lst:
        vals=[data[p][vk] for vk in variant_map]
        print(f"{p:8}" + "".join(f"{x:13.4g}" for x in vals))
