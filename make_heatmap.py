import openpyxl, glob, os, re

base = r"D:\我的文档\研究生\组会2026\7.8组会\AdaMao实验表\消融实验\指标模式"
files = sorted([f for f in glob.glob(os.path.join(base, "*.xlsx")) if not os.path.basename(f).startswith("~$")])
variant_map = {
    'REMO_new2_AdaMaO': 'Full', 'REMO_new2_AdaMaO_NoIndicator': 'NoInd',
    'REMO_new2_AdaMaO_Lite': 'Lite', 'REMO_new2_AdaMaO_PIEAOnlySelection': 'PIEAonly',
    'REMO_new2_AdaMaO_FixedIndicatorAlways': 'FixedInd',
}
variants = ['Full','NoInd','Lite','PIEAonly','FixedInd']
pat = re.compile(r'([-+]?\d+\.\d+e[+-]\d+)\s*\(([-+]?\d+\.\d+e[+-]\d+)\)\s*([+\-=])?')
data = {}; problems = []
for f in files:
    wb = openpyxl.load_workbook(f, data_only=True); ws = wb.active; header=[c.value for c in ws[1]]
    for r in range(2, ws.max_row+1):
        row=[ws.cell(row=r,column=c).value for c in range(1,ws.max_column+1)]
        p=row[0]
        if not isinstance(p,str) or not (p.startswith('DTLZ') or p.startswith('WFG')): continue
        problems.append(p)
        for ci,name in enumerate(header):
            if name is None: continue
            m=pat.search(str(row[ci])) if isinstance(row[ci],str) else None
            if m: data.setdefault(p,{})[name]=float(m.group(1))
problems=list(dict.fromkeys(problems))

# per-problem rank (1=best)
ranks={}; vals={}
for p in problems:
    v=[data[p][vk] for vk in variant_map]; vals[p]=v
    order=sorted(range(5),key=lambda i:v[i]); r=[0]*5; i=0
    while i<5:
        j=i
        while j<4 and v[order[j+1]]==v[order[i]]: j+=1
        ar=(i+j)/2+1
        for k in range(i,j+1): r[order[k]]=ar
        i=j+1
    ranks[p]=r

# average rank
avgr={v:0 for v in variants}
for p in problems:
    for idx,v in enumerate(variants): avgr[v]+=ranks[p][idx]
for v in variants: avgr[v]/=len(problems)

# color scale by rank 1..5  (green->yellow->red)
def color(rk):
    # rk in (1,5]; 1 best -> green, 5 worst -> red
    t=(rk-1)/4.0
    # green(120) to red(0) hue
    hue=int(120*(1-t))
    return f"hsl({hue},65%,{88-int(t*18)}%)"

disconn={'DTLZ7','WFG6'}
html=[]
html.append("""<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8">
<style>
body{font-family:-apple-system,"Segoe UI","Microsoft YaHei",sans-serif;background:#fff;color:#222;margin:24px;}
h2{font-size:18px;margin:0 0 4px} .sub{color:#666;font-size:12px;margin-bottom:16px}
table{border-collapse:collapse;font-size:12px}
th,td{border:1px solid #e0e0e0;padding:5px 7px;text-align:center}
th{background:#f5f5f5;font-weight:600}
td.prob{font-weight:600;text-align:left;background:#fafafa}
.barwrap{margin:18px 0}
.barrow{display:flex;align-items:center;margin:4px 0;font-size:12px}
.bname{width:70px}
.btrack{background:#eee;flex:1;height:18px;border-radius:3px;position:relative}
.bfill{height:18px;background:#4285f4;border-radius:3px}
.bval{width:60px;text-align:right;color:#444}
.legend{font-size:11px;color:#666;margin-top:8px}
.note{background:#fff8e1;border-left:4px solid #ffb300;padding:10px 12px;font-size:13px;margin:16px 0;border-radius:3px}
.good{background:#e8f5e9} .bad{background:#ffebee}
</style></head><body>""")

html.append("<h2>REMO_new2_AdaMaO 指标模式（PIEA）消融 — 逐问题 IGD 热力图</h2>")
html.append("<div class='sub'>5 个变体 × 16 问题（DTLZ1–7 + WFG1–9，全部 M=10，D=30/31，IGD 越小越好）。颜色：绿=该问题最佳(秩1)，红=最差(秩5)。数据来自 5 个 xlsx（均值±标准差一致，仅显著性控制算法不同）。</div>")

html.append("<table><tr><th>Problem</th><th>Full</th><th>NoInd</th><th>Lite</th><th>PIEAonly</th><th>FixedInd</th><th>该问题最佳</th></tr>")
for p in problems:
    cells=""
    best=min(vals[p]); bestv=[variants[i] for i,x in enumerate(vals[p]) if x==best]
    for idx,v in enumerate(variants):
        rk=ranks[p][idx]; x=vals[p][idx]
        cls=""
        if p in disconn and v in ('PIEAonly','FixedInd'): cls='bad'
        cells+=f"<td style='background:{color(rk)}'{(' class=\"'+cls+'\"') if cls else ''}>{x:.3g}<br><span style='font-size:10px;color:#555'>秩{rk:.1f}</span></td>"
    flag=" ⚠断开/退化PF" if p in disconn else ""
    cells+=f"<td style='font-size:11px'>{'/'.join(bestv)}{flag}</td>"
    html.append(f"<tr><td class='prob'>{p}</td>{cells}</tr>")
html.append("</table>")
html.append("<div class='legend'>⚠ = 断开/退化 PF（DTLZ7、WFG6）。注意 PIEAonly / FixedInd 在这两类问题上整行变红（崩塌），而在连通规则 PF 上多为绿色。</div>")

html.append("<div class='note'><b>关键规律：</b>指标模式（PIEAonly / FixedInd 的 indicator 部分）在<b>连通、规则 PF</b>（DTLZ1–5、WFG2、WFG9）上普遍更优（绿），但在<b>断开/退化 PF</b>（DTLZ7、WFG6）上<b>剧烈崩塌</b>——PIEAonly 在 DTLZ7 上 IGD=20.3（Full 仅 8.63，劣 2.35×）。Full 的「退化度≥0.45 才触发」机制正是防止这种崩塌的安全闸。</div>")

# average rank bar
html.append("<div class='barwrap'><h2 style='font-size:15px'>平均秩（越低越好，Friedman p=0.84 无显著差异）</h2>")
maxr=max(avgr.values())
for v in variants:
    w=avgr[v]/maxr*100
    html.append(f"<div class='barrow'><div class='bname'>{v}</div><div class='btrack'><div class='bfill' style='width:{w:.1f}%'></div></div><div class='bval'>{avgr[v]:.2f}</div></div>")
html.append("</div>")
html.append("</body></html>")

out=r"D:\PlatEMO-master\指标模式消融_热力图.html"
with open(out,"w",encoding="utf-8") as f: f.write("\n".join(html))
print("written",out)
print("avg ranks:",{v:round(avgr[v],3) for v in variants})
