import openpyxl, glob, os

base = r"D:\我的文档\研究生\组会2026\7.8组会\AdaMao实验表\消融实验\指标模式"
files = glob.glob(os.path.join(base, "*.xlsx"))

for f in sorted(files):
    if os.path.basename(f).startswith("~$"):
        continue
    print("="*80)
    print("FILE:", os.path.basename(f))
    wb = openpyxl.load_workbook(f, data_only=True)
    print("SHEETS:", wb.sheetnames)
    ws = wb.active
    print("DIM:", ws.dimensions, "max_row", ws.max_row, "max_col", ws.max_column)
    # print first 8 rows
    for r in range(1, min(ws.max_row, 8)+1):
        vals = []
        for c in range(1, ws.max_column+1):
            v = ws.cell(row=r, column=c).value
            vals.append(v)
        print(r, vals)
