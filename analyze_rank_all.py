import openpyxl, glob, os, re

base = r"D:\我的文档\研究生\组会2026\7.8组会\AdaMao实验表\消融实验\指标模式"
files = sorted([f for f in glob.glob(os.path.join(base, "*.xlsx")) if not os.path.basename(f).startswith("~$")])
pat = re.compile(r'([-+]?\d+\.\d+e[+-]\d+)\s*\(([-+]?\d+\.\d+e[+-]\d+)\)\s*([+\-=])?')

# all 11 algorithms we want to rank together
algos = ['REMO_new2_AdaMaO','REMO_new2_AdaMaO_NoIndicator','REMO_new2_AdaMaO_Lite',
         'REMO_new2_AdaMaO_PIEAOnlySelection','REMO_new2_AdaMaO_FixedIndicatorAlways',
         'REMO_new2','REMO','PIEA','MCEAD','REMO_new2_WFG10','R2AEA']
short = {'REMO_new2_AdaMaO':'Full','REMO_new2_AdaMaO_NoIndicator':'NoInd','REMO_new2_AdaMaO_Lite':'Lite',
         'REMO_new2_AdaMaO_PIEAOnlySelection':'PIEAonly','REMO_new2_AdaMaO_FixedIndicatorAlways':'FixedInd',
         'REMO_new2':'REMO_new2','REMO':'REMO','PIEA':'PIEA','MCEAD':'MCEAD','REMO_new2_WFG10':'WFG10','R2AEA':'R2AEA'}

data={}; problems=[]
for f in files:
    wb=openpyxl.load_workbook(f,data_only=True); ws=wb.active; header=[c.value for c in ws[1]]
    for r in range(2,ws.max_row+1):
        row=[ws.cell(row=r,column=c).value for c in range(1,ws.max_column+1)]
        p=row[0]
        if not isinstance(p,str) or not (p.startswith('DTLZ') or p.startswith('WFG')): continue
        problems.append(p)
        for ci,name in enumerate(header):
            if name is None: continue
            m=pat.search(str(row[ci])) if isinstance(row[ci],str) else None
            if m and name in algos: data.setdefault(p,{})[name]=float(m.group(1))
problems=list(dict.fromkeys(problems))

# average rank across ALL 11 algos, per problem
rank_sum={a:0.0 for a in algos}
for p in problems:
    present=[a for a in algos if a in data[p]]
    vals=[data[p][a] for a in present]
    order=sorted(range(len(present)),key=lambda i:vals[i])
    r=[0.0]*len(present); i=0
    while i<len(present):
        j=i
        while j<len(present)-1 and vals[order[j+1]]==vals[order[i]]: j+=1
        ar=(i+j)/2+1
        for k in range(i,j+1): r[order[k]]=ar
        i=j+1
    for idx,a in enumerate(present): rank_sum[a]+=r[idx]

N=len(problems)
print(f"=== Average rank across ALL {len(algos)} algorithms ({N} problems, lower=better) ===")
ordered=sorted(algos,key=lambda a: rank_sum[a]/N)
for a in ordered:
    tag="  <-- 本算法变体" if a.startswith('REMO_new2_AdaMaO') else ""
    print(f"{short[a]:12} avg_rank={rank_sum[a]/N:5.2f}{tag}")

# How do the 5 variants rank WITHIN the full field? (their rank position among 11)
print("\n=== 5 variant 在全场 11 算法中的名次（1=全场最佳）===")
full_avg={a:rank_sum[a]/N for a in algos}
for v in ['REMO_new2_AdaMaO','REMO_new2_AdaMaO_NoIndicator','REMO_new2_AdaMaO_Lite','REMO_new2_AdaMaO_PIEAOnlySelection','REMO_new2_AdaMaO_FixedIndicatorAlways']:
    pos=sorted(full_avg,key=lambda a:full_avg[a]).index(v)+1
    print(f"{short[v]:12} 全场第 {pos} / {len(algos)} 名, avg_rank={full_avg[v]:.2f}")
